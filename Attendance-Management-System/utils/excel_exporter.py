from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models.settings import CollegeSettings

def generate_excel_report(title, subtitle, headers, data_rows, summary_stats=None):
    """
    Generates styled Excel file (.xlsx) with dynamic College Branding in headers.
    Returns BytesIO object containing Excel spreadsheet bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    ws.views.sheetView[0].showGridLines = True

    settings = CollegeSettings.get_settings()

    # Fonts & Fills
    college_title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
    sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    report_title_font = Font(name="Calibri", size=13, bold=True, color="0F172A")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="1E293B")

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. College Header
    ws.cell(row=1, column=1, value=settings.college_name.upper()).font = college_title_font
    ws.cell(row=2, column=1, value=f"{settings.address} | Contact: {settings.contact_number} | Email: {settings.email_address}").font = sub_font
    ws.cell(row=3, column=1, value=f"Academic Year: {settings.academic_year} | Principal: {settings.principal_name}").font = sub_font
    
    # 2. Report Title & Subtitle
    ws.cell(row=5, column=1, value=title).font = report_title_font
    current_row = 6
    if subtitle:
        ws.cell(row=current_row, column=1, value=subtitle).font = sub_font
        current_row += 1

    current_row += 1

    # 3. Summary Stats (if provided)
    if summary_stats:
        for k, v in summary_stats.items():
            cell_k = ws.cell(row=current_row, column=1, value=f"{k}:")
            cell_v = ws.cell(row=current_row, column=2, value=str(v))
            cell_k.font = Font(name="Calibri", size=10, bold=True)
            cell_v.font = Font(name="Calibri", size=10)
            current_row += 1
        current_row += 1

    # 4. Data Headers
    header_row_idx = current_row
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws.row_dimensions[header_row_idx].height = 24
    current_row += 1

    # 5. Data Rows
    for row_data in data_rows:
        ws.row_dimensions[current_row].height = 20
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=cell_value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
            if (current_row - header_row_idx) % 2 == 0:
                cell.fill = alt_row_fill
        current_row += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
